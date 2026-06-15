import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

wb = openpyxl.Workbook()

hdr_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
hdr_fill = PatternFill('solid', fgColor='1C3D5A')
cat_font = Font(bold=True, color='1C3D5A', size=10, name='Arial')
cat_fill = PatternFill('solid', fgColor='E9EFF5')
cell_font = Font(size=10, name='Arial')
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

def add_header(ws, row, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def add_cat(ws, row, title, ncols):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = cat_font
    cell.fill = cat_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell.border = thin_border

def add_row(ws, row, vals):
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = cell_font
        cell.border = thin_border

with open(r'C:\Users\Senko\code\hidrocalc\src\components\ListaMaterialesSIMEX.tsx', 'r', encoding='utf-8') as f:
    src = f.read()

# Sheet 1: Kit Bridas
ws = wb.active
ws.title = 'Kit Bridas'
for col, w in [('A',18),('B',20),('C',22),('D',22),('E',18),('F',22),('G',8)]:
    ws.column_dimensions[col].width = w
add_header(ws, 1, ['DN|Material', 'ABU SKU', 'Extremidad SKU', 'Gibault SKU', 'Empaque SKU', 'Tornillo SKU', 'Bolts'])

kit_pattern = r"'([^']+)':\s*\{([^}]+)\}"
row = 2
for m in re.finditer(kit_pattern, src[:12000]):
    key = m.group(1)
    data = m.group(2)
    abu = re.search(r'"a":\s*"([^"]+)"', data)
    ext = re.search(r'"e":\s*"([^"]+)"', data)
    gib = re.search(r'"g":\s*"([^"]+)"', data)
    emp = re.search(r'"em":\s*"([^"]+)"', data)
    tor = re.search(r'"t":\s*"([^"]+)"', data)
    bol = re.search(r'"b":\s*(\d+)', data)
    add_row(ws, row, [
        key,
        abu.group(1) if abu else '',
        ext.group(1) if ext else '',
        gib.group(1) if gib else '',
        emp.group(1) if emp else '',
        tor.group(1) if tor else '',
        int(bol.group(1)) if bol else ''
    ])
    row += 1

# Sheet 2: Conexiones
ws2 = wb.create_sheet('Conexiones')
for col, w in [('A',15),('B',10),('C',10),('D',20),('E',8)]:
    ws2.column_dimensions[col].width = w
add_header(ws2, 1, ['Tipo', 'DN1', 'DN2', 'SKU', 'Bridas'])

conn_pattern = r'\{"f":\s*"([^"]+)",\s*"d1":\s*"([^"]+)",\s*"d2":\s*"([^"]+)",\s*"sk":\s*"([^"]+)",\s*"br":\s*(\d+)\}'
row2 = 2
for m in re.finditer(conn_pattern, src):
    add_row(ws2, row2, [m.group(1), m.group(2), m.group(3), m.group(4), int(m.group(5))])
    row2 += 1

# Sheet 3: Valvulas
ws3 = wb.create_sheet('Valvulas')
for col, w in [('A',35),('B',8),('C',20)]:
    ws3.column_dimensions[col].width = w
add_header(ws3, 1, ['Tipo', 'DN', 'SKU'])

valv_data = [
    ('vcg-r', 'V. Compuerta Resilente C515 Sigma Flow'),
    ('vcg-b', 'V. Compuerta Bronce C500 Sigma Flow'),
    ('vmb-c', 'V. Mariposa C504 Sigma Flow'),
    ('vmb-dex', 'V. Mariposa Doble Exc. Sigma Flow'),
    ('vmb-w', 'V. Mariposa Wafer Sigma Flow'),
]
row3 = 2
for vkey, vlabel in valv_data:
    pat = r"'" + re.escape(vkey) + r"':\s*\{([^}]+)\}"
    vm = re.search(pat, src)
    if vm:
        pairs = re.findall(r"'([^']+)':\s*\"([^\"]+)\"", vm.group(1))
        for dn, sku in pairs:
            add_row(ws3, row3, [vlabel, dn, sku])
            row3 += 1

# Sheet 4: Tapas y Carretes
ws4 = wb.create_sheet('Tapas y Carretes')
for col, w in [('A',25),('B',8),('C',20)]:
    ws4.column_dimensions[col].width = w
add_header(ws4, 1, ['Tipo', 'DN', 'SKU'])

row4 = 2
add_cat(ws4, row4, 'TAPAS CIEGAS HD Sigma', 3); row4 += 1
tapa_pattern = r'"(\d+[\\]?"|\d+[^"]*?)":\s*"(CI-BCF-[^"]+)"'
for m in re.finditer(tapa_pattern, src):
    dn = m.group(1).replace('\\', '')
    add_row(ws4, row4, ['Tapa Ciega HD Sigma', dn, m.group(2)])
    row4 += 1

add_cat(ws4, row4, 'CARRETES DE DESMONTAJE Sigma Flow', 3); row4 += 1
cdm_pattern = r"'([^']+)':\s*\"(CI-CDM-[^\"]+)\""
for m in re.finditer(cdm_pattern, src):
    add_row(ws4, row4, ['Carrete de Desmontaje Sigma Flow', m.group(1), m.group(2)])
    row4 += 1

# Sheet 5: Checks
ws5 = wb.create_sheet('Checks')
for col, w in [('A',35),('B',8),('C',20)]:
    ws5.column_dimensions[col].width = w
add_header(ws5, 1, ['Tipo', 'DN', 'SKU'])

row5 = 2
add_cat(ws5, row5, 'CHECK RESILENTE C508 (2"-16")', 3); row5 += 1
for d in ['2','3','4','6','8','10','12','14','16']:
    add_row(ws5, row5, ['Check Resilente C508 Sigma Flow', d+'"', f'VI-CHK-R{d}']); row5 += 1

add_cat(ws5, row5, 'CHECK COMPUERTA BRONCE (18"+)', 3); row5 += 1
for d in ['18','20','24','30','36']:
    add_row(ws5, row5, ['Check Compuerta Bronce Sigma Flow', d+'"', f'VI-CHK-{d}']); row5 += 1

add_cat(ws5, row5, 'DUO CHECK WAFER', 3); row5 += 1
for d in ['2','3','4','6','8','10','12']:
    add_row(ws5, row5, ['Duo Check Wafer Sigma Flow', d+'"', f'VI-DCK-{d}']); row5 += 1

# Sheet 6: ABU Descripciones
ws6 = wb.create_sheet('ABU Descripciones')
ws6.column_dimensions['A'].width = 22
ws6.column_dimensions['B'].width = 65
add_header(ws6, 1, ['SKU', 'Descripcion'])

abu_desc = {
    'CI-ABU-2': 'Adaptador de Brida Universal de 2" de 59 a 72 mm Sigma Flow',
    'CI-ABU-24860': 'Adaptador de Brida Universal de 2" de 48 a 60 mm Sigma Flow',
    'CI-ABU-257285': 'Adaptador de Brida Universal de 2.5" de 72 a 85 mm Sigma Flow',
    'CI-ABU-3': 'Adaptador de Brida Universal de 3" de 88 a 103 mm Sigma Flow',
    'CI-ABU-496116': 'Adaptador de Brida Universal de 4" de 96 a 116 mm Sigma Flow',
    'CI-ABU-4109130': 'Adaptador de Brida Universal de 4" de 109 a 130 mm Sigma Flow',
    'CI-ABU-4107135': 'Adaptador de Brida Universal de 4" de 107 a 135 mm Sigma Flow',
    'CI-ABU-6159184': 'Adaptador de Brida Universal de 6" de 159 a 184 mm Sigma Flow',
    'CI-ABU-8I': 'Adaptador de Brida Universal de 8" Ingles Sigma Flow',
    'CI-ABU-8M': 'Adaptador de Brida Universal de 8" Metrico Sigma Flow',
    'CI-ABU-8214249': 'Adaptador de Brida Universal de 8" de 214 a 249 mm Sigma Flow',
    'CI-ABU-10AC': 'Adaptador de Brida Universal de 10" de 272 a 289 mm Sigma Flow',
    'CI-ABU-10272308': 'Adaptador de Brida Universal de 10" de 272 a 308 mm Sigma Flow',
    'CI-ABU-10245267': 'Adaptador de Brida Universal de 10" de 245 a 267 mm Sigma Flow',
    'CI-ABU-12': 'Adaptador de Brida Universal de 12" de 315 a 332 mm Sigma Flow',
    'CI-ABU-12322342': 'Adaptador de Brida Universal de 12" de 322 a 342 mm Sigma Flow',
    'CI-ABU-12324365': 'Adaptador de Brida Universal de 12" de 324 a 365 mm Sigma Flow',
    'CI-ABU-14': 'Adaptador de Brida Universal de 14" de 351 a 378 mm Sigma Flow',
    'CI-ABU-14374391': 'Adaptador de Brida Universal de 14" de 374 a 391 mm Sigma Flow',
    'CI-ABU-16': 'Adaptador de Brida Universal de 16" de 390 a 410 mm Sigma Flow',
    'CI-ABU-16425442': 'Adaptador de Brida Universal de 16" de 425 a 442 mm Sigma Flow',
    'CI-ABU-16390435': 'Adaptador de Brida Universal de 16" de 390 a 435 mm Sigma Flow',
    'CI-ABU-18445472': 'Adaptador de Brida Universal de 18" de 445 a 472 mm Sigma Flow',
    'CI-ABU-18480510': 'Adaptador de Brida Universal de 18" de 480 a 510 mm Sigma Flow',
    'CI-ABU-20500532': 'Adaptador de Brida Universal de 20" de 500 a 532 mm Sigma Flow',
    'CI-ABU-20527544': 'Adaptador de Brida Universal de 20" de 527 a 544 mm Sigma Flow',
    'CI-ABU-24': 'Adaptador de Brida Universal de 24" de 608 a 636 mm Sigma Flow',
    'CI-ABU-24645680': 'Adaptador de Brida Universal de 24" de 645 a 680 mm Sigma Flow',
    'CI-ABP-EAD3': 'Adaptador de Brida Universal de 3" PEAD Sigma Flow',
    'CI-ABP-EAD4': 'Adaptador de Brida Universal de 4" PEAD Sigma Flow',
    'CI-ABP-EAD6': 'Adaptador de Brida Universal de 6" PEAD Sigma Flow',
    'CI-ABP-EAD8': 'Adaptador de Brida Universal de 8" PEAD Sigma Flow',
    'CI-ABP-EAD10': 'Adaptador de Brida Universal de 10" PEAD Sigma Flow',
    'CI-ABP-EAD12': 'Adaptador de Brida Universal de 12" PEAD Sigma Flow',
}
row6 = 2
for sku, desc in abu_desc.items():
    add_row(ws6, row6, [sku, desc])
    row6 += 1

# Sheet 7: Tornillos
ws7 = wb.create_sheet('Tornillos')
ws7.column_dimensions['A'].width = 22
ws7.column_dimensions['B'].width = 35
add_header(ws7, 1, ['SKU', 'Descripcion'])
tor_desc = {
    'DN-TOR-5/821/2': 'Tornillo 5/8" x 2-1/2"',
    'DN-TOR-5/83': 'Tornillo 5/8" x 3"',
    'DN-TOR-3/431/2': 'Tornillo 3/4" x 3-1/2"',
    'DN-TOR-7/84': 'Tornillo 7/8" x 4"',
    'DN-TOR-141/2': 'Tornillo 1" x 4-1/2"',
    'DN-TOR-11/85': 'Tornillo 1-1/8" x 5"',
    'DN-TOR-11/451/2': 'Tornillo 1-1/4" x 5-1/2"',
    'DN-TOR-11/46': 'Tornillo 1-1/4" x 6"',
    'DN-TOR-11/27': 'Tornillo 1-1/2" x 7"',
}
row7 = 2
for sku, desc in tor_desc.items():
    add_row(ws7, row7, [sku, desc])
    row7 += 1

out = r'C:\Users\Senko\code\hidrocalc\Catalogo_SIMEX_HidroCalc.xlsx'
wb.save(out)
print(f'OK: {out}')
